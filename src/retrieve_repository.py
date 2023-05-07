from datetime import date
from enum import Enum
import time
import pyexcel as p
import openpyxl
from os.path import exists
from src.ids import create_id
from src.movements import Movement, parse_numbers


def change_file_extension(file: str):
    path = file + ".xlsx"
    if not exists(path):
        p.save_book_as(file_name=file + ".xls", dest_file_name=path)
    return path


def get_active_sheet(path: str):
    wb_obj = openpyxl.load_workbook(path)
    return wb_obj.active


def get_account_id(sheet_obj):
    return create_id(sheet_obj.cell(row=2, column=4).value)


def get_seet(file):
    path = change_file_extension(file)
    return get_active_sheet(path)


def get_movements(sheet):
    movements = []
    for row_index in range(7, sheet.max_row):
        row = sheet[row_index]
        datetime = date(
            int(row[0].value.year), int(row[0].value.month), int(row[0].value.day)
        )
        timestamp = time.mktime(datetime.timetuple())
        amount = row[6].value
        movements.append(
            Movement(
                id=create_id(str(timestamp) + row[3].value + str(amount)),
                comment=row[4].value or row[3].value,
                description=row[3].value,
                category="",
                amount=amount,
                day=datetime.day,
                month=datetime.month,
                year=datetime.year,
            )
        )
    return movements


def get_movements_credit_card(sheet):
    movements = []
    for row_index in range(7, sheet.max_row):
        row = sheet[row_index]
        datetime = date(
            int(row[0].value.year), int(row[0].value.month), int(row[0].value.day)
        )
        timestamp = time.mktime(datetime.timetuple())
        amount = row[7].value
        movements.append(
            Movement(
                id=create_id(str(timestamp) + row[3].value + str(amount)),
                comment=row[4].value or row[3].value,
                description=row[3].value,
                category="",
                amount=amount,
                day=datetime.day,
                month=datetime.month,
                year=datetime.year,
            )
        )

    return movements


def get_movements_factory(account_type, sheet):
    if account_type == AccountType.REGULAR or account_type == AccountType.SAVING:
        return get_movements(sheet)

    if account_type == AccountType.CREDIT_CARD:
        print(account_type)
        return get_movements_credit_card(sheet)


class AccountType(Enum):
    REGULAR = "regular"
    SAVING = "saving"
    CREDIT_CARD = "credit"
    INVESTMENT = "investment"
